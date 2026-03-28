"""Export leads to CSV, Excel, and Google Sheets."""

import os
from datetime import datetime

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from src import config


# ── CSV Export ───���───────────────────────��─────────────────────────────

def to_csv(leads: list[dict], industry: str, city: str) -> str:
    """
    Save leads to a CSV file in the leads/ folder.

    Returns:
        Full path to the saved CSV file.
    """
    df = _leads_to_dataframe(leads)
    date_str = datetime.now().strftime("%Y-%m-%d")

    # Clean filename
    industry_clean = industry.lower().replace(" ", "_")
    city_clean = city.lower().replace(" ", "_").replace(",", "")
    filename = f"{industry_clean}_{city_clean}_{date_str}.csv"
    filepath = os.path.join(config.LEADS_DIR, filename)

    df.to_csv(filepath, index=False)
    print(f"[export] Saved {len(df)} leads to {filepath}")
    return filepath


# ── Excel Export ────���──────────────────────────────────────────────────

HEADER_LABELS = {
    "name": "Business Name",
    "phone": "Phone",
    "email": "Email",
    "website": "Website",
    "address": "Address",
    "city": "City",
    "country": "Country",
    "rating": "Rating",
    "reviews_count": "Reviews",
    "category": "Category",
    "opening_hours": "Opening Hours",
    "google_maps_url": "Google Maps Link",
}


def to_excel(sheets: dict[str, list[dict]], filename: str) -> str:
    """
    Create a professional Excel workbook with multiple sheets.

    Args:
        sheets: Dict of {sheet_name: leads_list}
                e.g. {"Dentists - Dubai": [...], "Construction - Houston": [...]}
        filename: Output filename (without path)

    Returns:
        Full path to the saved Excel file.
    """
    filepath = os.path.join(config.LEADS_DIR, filename)

    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        for sheet_name, leads in sheets.items():
            df = _leads_to_dataframe(leads)
            # Rename columns to professional headers
            df.columns = [HEADER_LABELS.get(c, c) for c in df.columns]
            df.to_excel(writer, sheet_name=sheet_name[:31], index=False)

        # Style each sheet
        wb = writer.book
        for sheet_name in wb.sheetnames:
            _style_sheet(wb[sheet_name])

    print(f"[export] Saved Excel: {filepath}")
    return filepath


def _style_sheet(ws):
    """Apply professional formatting to an Excel worksheet."""
    # Colors
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    data_font = Font(name="Calibri", size=10)
    alt_fill = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")
    thin_border = Border(
        bottom=Side(style="thin", color="D9E2EC"),
    )

    # Style header row
    for col_idx, cell in enumerate(ws[1], 1):
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Style data rows
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), 2):
        for cell in row:
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=False)
        # Alternate row shading
        if row_idx % 2 == 0:
            for cell in row:
                cell.fill = alt_fill

    # Auto-fit column widths
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                val = str(cell.value) if cell.value else ""
                max_len = max(max_len, len(val))
        # Cap width to avoid absurdly wide columns
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    # Freeze header row
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = ws.dimensions


# ── Google Sheets Export ─────���─────────────────────────────────────────

def to_google_sheets(leads: list[dict], sheet_name: str | None = None) -> str:
    """
    Push leads to a Google Sheets spreadsheet.

    Args:
        leads: List of lead dicts
        sheet_name: Tab name inside the spreadsheet. Defaults to "Leads".

    Returns:
        URL of the spreadsheet.
    """
    import gspread
    from google.oauth2.service_account import Credentials

    creds = Credentials.from_service_account_file(
        config.GOOGLE_SHEETS_CREDENTIALS_PATH,
        scopes=["https://www.googleapis.com/auth/spreadsheets"],
    )
    gc = gspread.authorize(creds)

    spreadsheet = gc.open(config.GOOGLE_SHEETS_SPREADSHEET_NAME)
    tab_name = sheet_name or "Leads"

    # Create or get worksheet
    try:
        worksheet = spreadsheet.worksheet(tab_name)
    except gspread.exceptions.WorksheetNotFound:
        worksheet = spreadsheet.add_worksheet(title=tab_name, rows=1000, cols=20)

    df = _leads_to_dataframe(leads)

    # Clear and write
    worksheet.clear()
    worksheet.update([df.columns.values.tolist()] + df.values.tolist())

    url = spreadsheet.url
    print(f"[export] Pushed {len(df)} leads to Google Sheets: {url}")
    return url


# ── Data Cleaning ──────────────────────────────────────────────────────

def clean_leads(leads: list[dict], min_rating: float | None = None) -> list[dict]:
    """
    Deduplicate and clean the lead list.

    - Removes duplicates (same phone or same website)
    - Removes leads with no phone AND no email
    - Optionally filters by minimum rating
    - Sorts by rating (highest first)
    """
    df = _leads_to_dataframe(leads)
    initial = len(df)

    # Dedup by phone (if present)
    phone_mask = df["phone"].astype(str).str.strip().ne("")
    df_with_phone = df[phone_mask].drop_duplicates(subset=["phone"], keep="first")
    df_no_phone = df[~phone_mask]

    # Dedup by website for those without phone
    website_mask = df_no_phone["website"].astype(str).str.strip().ne("")
    df_with_website = df_no_phone[website_mask].drop_duplicates(subset=["website"], keep="first")
    df_neither = df_no_phone[~website_mask]

    df = pd.concat([df_with_phone, df_with_website, df_neither], ignore_index=True)

    # Remove leads with no phone AND no email
    df = df[
        (df["phone"].astype(str).str.strip().ne(""))
        | (df["email"].astype(str).str.strip().ne(""))
    ]

    # Filter by rating
    if min_rating is not None:
        df["rating"] = pd.to_numeric(df["rating"], errors="coerce")
        df = df[df["rating"] >= min_rating]

    # Sort by rating descending
    df["rating"] = pd.to_numeric(df["rating"], errors="coerce")
    df = df.sort_values("rating", ascending=False).reset_index(drop=True)

    cleaned = len(df)
    print(f"[clean] {initial} -> {cleaned} leads (removed {initial - cleaned} duplicates/empty)")

    return df.to_dict("records")


# ── Helpers ────────────────────────────────────────────────────────────

EXPORT_COLUMNS = [
    "name", "phone", "email", "website", "address", "city", "country",
    "rating", "reviews_count", "category", "opening_hours", "google_maps_url",
]


def _leads_to_dataframe(leads: list[dict]) -> pd.DataFrame:
    """Convert lead dicts to a clean DataFrame with standard columns."""
    df = pd.DataFrame(leads)
    # Only keep export columns that exist
    cols = [c for c in EXPORT_COLUMNS if c in df.columns]
    return df[cols]


def print_summary(leads: list[dict]) -> None:
    """Print a quick summary of the lead list."""
    total = len(leads)
    with_phone = sum(1 for l in leads if l.get("phone"))
    with_email = sum(1 for l in leads if l.get("email"))
    with_website = sum(1 for l in leads if l.get("website"))

    print("\n" + "=" * 50)
    print(f"  LEAD SCRAPING RESULTS")
    print("=" * 50)
    print(f"  Total leads:    {total}")
    print(f"  With phone:     {with_phone} ({_pct(with_phone, total)})")
    print(f"  With email:     {with_email} ({_pct(with_email, total)})")
    print(f"  With website:   {with_website} ({_pct(with_website, total)})")
    print("=" * 50 + "\n")


def _pct(part: int, total: int) -> str:
    if total == 0:
        return "0%"
    return f"{round(part / total * 100)}%"
